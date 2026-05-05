from flask import Blueprint, render_template, request, flash, redirect, url_for, Response, jsonify
from flask_login import login_required, current_user
from models import db, Student, Payment, Attendance, Expenditure
from datetime import datetime, timedelta
from sqlalchemy import or_, func
import random
import string
import csv
from io import StringIO

main_routes = Blueprint('main', __name__)

def _new_cash_transaction_number(student_id: int) -> str:
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"CASH-{student_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{suffix}"

def get_sessions_data():
    """Helper function to get sessions data with student counts"""
    sessions = db.session.query(
        Student.session,
        db.func.count(Student.id).label('total_students'),
        db.func.sum(db.case((Student.active == True, 1), else_=0)).label('active_students')
    ).group_by(Student.session).order_by(Student.session.desc()).all()
    
    return [{
        'session': session,
        'total_students': total,
        'active_students': active or 0
    } for session, total, active in sessions if session]  # Filter out None sessions

@main_routes.route('/')
@login_required
def index():
    from flask_login import current_user
    print(f"[DEBUG] Dashboard: is_authenticated={current_user.is_authenticated}, user={getattr(current_user, 'username', None)}")
    
    # Get sessions data
    sessions_data = get_sessions_data()
    
    # Get current session (latest with active students)
    current_session = db.session.query(Student.session).filter(Student.active == True).order_by(Student.session.desc()).distinct().first()
    current_session = current_session[0] if current_session else None
    
    print(f"[DEBUG] Current session detected: {current_session}")
    
    # Calculate stats for current session only
    total_students = Student.query.filter_by(active=True).count()
    
    # Count distinct active students with ANY cleared/partial payment of that type in current session
    passport_query = db.session.query(Student.id).distinct().join(Payment).filter(
        Payment.payment_type == 'Passport Fee',
        Student.active == True,
        Payment.status.in_(['cleared', 'partial'])
    )
    if current_session:
        passport_query = passport_query.filter(Student.session == current_session)
    passport_fee_count = passport_query.count()
    
    graduation_query = db.session.query(Student.id).distinct().join(Payment).filter(
        Payment.payment_type == 'Graduation Fee',
        Student.active == True,
        Payment.status.in_(['cleared', 'partial'])
    )
    if current_session:
        graduation_query = graduation_query.filter(Student.session == current_session)
    graduation_fee_count = graduation_query.count()
    
    # Calculate total revenue from all cleared payments (keep all-time)
    total_revenue = db.session.query(db.func.sum(Payment.amount)).filter(Payment.status == 'cleared').scalar() or 0
    
    # Calculate graduated students for current session
    graduated_query = db.session.query(Student).join(Payment).filter(
        Payment.payment_type == 'Graduation Fee',
        Payment.status == 'cleared',
        Student.active == True
    )
    if current_session:
        graduated_query = graduated_query.filter(Student.session == current_session)
    total_graduated = graduated_query.group_by(Student.id).having(db.func.sum(Payment.amount) >= 1000).count()

    # Attendance trend (last 14 days): list of {date, present}
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=13)
    present_by_date = dict(
        db.session.query(Attendance.date, db.func.count(Attendance.id))
        .filter(Attendance.date >= start_date, Attendance.date <= end_date)
        .filter(Attendance.status == 'present')
        .group_by(Attendance.date)
        .all()
    )
    trend_data = []
    d = start_date
    while d <= end_date:
        trend_data.append({"date": d.isoformat(), "present": int(present_by_date.get(d, 0))})
        d += timedelta(days=1)
    
    return render_template(
        'dark_dashboard.html',
        total_students=total_students,
        passport_fee_count=passport_fee_count,
        graduation_fee_count=graduation_fee_count,
        total_graduated=total_graduated,
        total_revenue=total_revenue,
        sessions_data=sessions_data,
        trend_data=trend_data,
    )

@main_routes.route('/students')
@login_required
def search_students():
    page = request.args.get('page', 1, type=int)
    search = (request.args.get('search') or '').strip()
    show_inactive = (request.args.get('show_inactive') == 'true')

    q = Student.query
    if not show_inactive:
        q = q.filter(Student.active == True)

    if search:
        # search by name, numeric ID, or admission number
        filters = [Student.name.ilike(f'%{search}%'), Student.admission_number.ilike(f'%{search}%')]
        if search.isdigit():
            filters.append(Student.id == int(search))
        q = q.filter(or_(*filters))

    q = q.order_by(Student.id.desc())
    pagination = q.paginate(page=page, per_page=12, error_out=False)
    students = pagination.items

    # Attach computed fields used by the template
    if students:
        student_ids = [s.id for s in students]
        totals = dict(
            db.session.query(Payment.student_id, func.coalesce(func.sum(Payment.amount), 0.0))
            .filter(Payment.student_id.in_(student_ids))
            .group_by(Payment.student_id)
            .all()
        )
        last_dates = dict(
            db.session.query(Payment.student_id, func.max(Payment.date))
            .filter(Payment.student_id.in_(student_ids))
            .group_by(Payment.student_id)
            .all()
        )
        for s in students:
            s.total_paid = float(totals.get(s.id, 0.0) or 0.0)
            s.last_payment_date = last_dates.get(s.id)

    return render_template('search_students.html', students=students, pagination=pagination)

@main_routes.route('/manage_payments/<int:student_id>', methods=['GET', 'POST'])
@login_required
def manage_payments(student_id):
    student = Student.query.get_or_404(student_id)

    if request.method == 'POST':
        payment_method = (request.form.get('payment_method') or '').strip().lower()
        payment_type = (request.form.get('payment_type') or '').strip()
        amount = request.form.get('amount', type=float)
        transaction_number = (request.form.get('transaction_number') or '').strip()

        if not payment_method or payment_method not in ['cash', 'mpesa']:
            flash('Please select a valid payment method.', 'error')
            return redirect(url_for('main.manage_payments', student_id=student.id))

        if not payment_type:
            flash('Please select a payment type.', 'error')
            return redirect(url_for('main.manage_payments', student_id=student.id))

        if amount is None or amount <= 0:
            flash('Please enter a valid amount.', 'error')
            return redirect(url_for('main.manage_payments', student_id=student.id))

        # Transaction number is required in DB; generate one for cash if not provided
        if payment_method == 'cash' and not transaction_number:
            transaction_number = _new_cash_transaction_number(student.id)
        elif payment_method == 'mpesa' and not transaction_number:
            flash('Please enter the M-PESA transaction number.', 'error')
            return redirect(url_for('main.manage_payments', student_id=student.id))

        total_fee = 1500.0
        if payment_type == 'Graduation Fee':
            total_fee = 1000.0
        elif payment_type == 'Passport Fee':
            total_fee = 500.0

        payment = Payment(
            student_id=student.id,
            transaction_number=transaction_number,
            amount=float(amount),
            payment_type=payment_type,
            payment_method=payment_method,
            status='pending',
            total_fee=total_fee,
            year=datetime.now().strftime('%Y'),
            session=student.session,
        )
        db.session.add(payment)
        db.session.commit()

        # Update status based on cumulative paid for this type
        try:
            payment.update_status()
        except Exception:
            # don't fail the flow if status calc fails
            pass

        flash('Payment added successfully.', 'success')
        return redirect(url_for('main.manage_payments', student_id=student.id))

    payments = Payment.query.filter_by(student_id=student.id).order_by(Payment.date.desc()).all()
    return render_template('manage_payments.html', student=student, payments=payments)

@main_routes.route('/receipt/<int:payment_id>')
@login_required
def view_receipt(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    student = Student.query.get_or_404(payment.student_id)
    return render_template('receipt.html', payment=payment, student=student)

@main_routes.route('/delete_payment/<int:payment_id>', methods=['POST'])
@login_required
def delete_payment(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    student_id = payment.student_id
    db.session.delete(payment)
    db.session.commit()
    flash('Payment deleted.', 'success')
    return redirect(url_for('main.manage_payments', student_id=student_id))

@main_routes.route('/remove_duplicates/<int:student_id>', methods=['POST'])
@login_required
def remove_duplicates(student_id):
    try:
        removed = Payment.remove_duplicates(student_id=student_id, dry_run=False)
        flash(f'Removed {removed} duplicate payment(s).', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Failed to remove duplicates: {str(e)}', 'error')
    return redirect(url_for('main.manage_payments', student_id=student_id))

@main_routes.route('/print_student_qr/<int:student_id>')
@login_required
def print_student_qr(student_id):
    student = Student.query.get_or_404(student_id)
    return render_template('print_qr.html', student=student)

@main_routes.route('/expenditure/add', methods=['GET', 'POST'])
@login_required
def add_expenditure():
    edit_id = request.args.get('edit_id', type=int)
    exp = Expenditure.query.get(edit_id) if edit_id else None

    if request.method == 'POST':
        date_str = (request.form.get('date') or '').strip()
        description = (request.form.get('description') or '').strip()
        amount = request.form.get('amount', type=float)

        if not date_str or not description or amount is None:
            flash('All fields are required.', 'error')
            return redirect(request.url)

        try:
            exp_date = datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            flash('Invalid date format.', 'error')
            return redirect(request.url)

        if exp is None:
            exp = Expenditure()
            db.session.add(exp)

        exp.date = exp_date
        exp.description = description
        exp.amount = float(amount)
        exp.created_by = getattr(current_user, 'id', None)

        db.session.commit()
        flash('Expenditure saved.', 'success')
        return redirect(url_for('main.expenditure_report'))

    total_expenditure = db.session.query(func.coalesce(func.sum(Expenditure.amount), 0.0)).scalar() or 0.0
    expenditure_count = db.session.query(func.count(Expenditure.id)).scalar() or 0
    date_default = datetime.now().strftime('%Y-%m-%d')
    return render_template(
        'add_expenditure.html',
        expenditure=exp,
        total_expenditure=total_expenditure,
        expenditure_count=expenditure_count,
        date_default=date_default,
    )

@main_routes.route('/expenditure/report')
@login_required
def expenditure_report():
    all_expenditures = Expenditure.query.order_by(Expenditure.date.desc()).all()
    total_expenditure = sum((e.amount or 0) for e in all_expenditures)
    expenditure_count = len(all_expenditures)
    avg_expenditure = (total_expenditure / expenditure_count) if expenditure_count else 0
    recent_expenditures = all_expenditures[:8]

    # Simple monthly aggregation for the chart
    monthly_rows = (
        db.session.query(func.strftime('%Y-%m', Expenditure.date).label('month'), func.sum(Expenditure.amount).label('total_amount'))
        .group_by('month')
        .order_by('month')
        .all()
    )
    expenditure_trends = [{'month': m, 'total_amount': float(v or 0)} for (m, v) in monthly_rows]

    return render_template(
        'expenditure_report.html',
        all_expenditures=all_expenditures,
        recent_expenditures=recent_expenditures,
        total_expenditure=total_expenditure,
        expenditure_count=expenditure_count,
        avg_expenditure=avg_expenditure,
        expenditure_trends=expenditure_trends,
    )

@main_routes.route('/expenditure/print/<int:id>')
@login_required
def expenditure_pdf(id):
    expenditure = Expenditure.query.get_or_404(id)
    print_date = datetime.now().strftime('%d %B %Y')
    return render_template('expenditure_print.html', expenditure=expenditure, print_date=print_date)

@main_routes.route('/student_payment_status', methods=['GET', 'POST'])
@login_required
def student_payment_status():
    sessions = db.session.query(Student.session).filter(Student.session.isnot(None)).distinct().all()
    sessions = sorted(set([s[0] for s in sessions if s[0] and str(s[0]).strip() != '']))
    # Fetch classes from Student model
    classes = db.session.query(Student.class_name).filter(Student.class_name.isnot(None)).distinct().all()
    classes = sorted(set([c[0] for c in classes if c[0] and str(c[0]).strip() != '']))
    classes.insert(0, "All Classes")
    payment_type = request.args.get("payment_type")
    report_data = None
    selected_class = None
    selected_session = None
    selected_payment_type = None
    total_enrolled = 0
    total_graduated = 0

    if payment_type:
        students = Student.query.filter_by(active=True).order_by(Student.class_name, Student.name).all()
        total_enrolled = len(students)
        total_cleared = 0
        for student in students:
            all_payments = student.payments
            passport_paid = sum(p.amount for p in all_payments if p.payment_type == 'Passport Fee')
            graduation_paid = sum(p.amount for p in all_payments if p.payment_type == 'Graduation Fee')
            total_paid = sum(p.amount for p in all_payments)
            student.payment_details = {
                'passport': passport_paid,
                'graduation': graduation_paid,
                'total': total_paid
            }
            payments = [p for p in all_payments if p.payment_type == payment_type]
            summary_paid = sum(p.amount for p in payments)
            total_fee = next((p.total_fee for p in payments), 1500.0)
            status = "cleared" if summary_paid >= total_fee else "partial" if summary_paid > 0 else "pending"
            student.payment_summary = {"total_paid": summary_paid, "total_fee": total_fee, "status": status}
            if status == "cleared":
                total_cleared += 1
        report_data = students
        selected_payment_type = payment_type
        total_graduated = total_cleared  # Reuse for cleared count
    
    if request.method == 'POST':
        selected_session = request.form.get('session')
        selected_class = request.form.get('class')
        if selected_session:
            if selected_class == "All Classes":
                students = Student.query.filter_by(session=selected_session).order_by(Student.class_name, Student.name).all()
                total_enrolled = Student.query.filter_by(session=selected_session).count()
            else:
                students = Student.query.filter_by(session=selected_session, class_name=selected_class).order_by(Student.name).all()
                total_enrolled = Student.query.filter_by(session=selected_session, class_name=selected_class).count()
            
            # Calculate payment details and graduated count
            for student in students:
                all_payments = student.payments
                passport_paid = sum(p.amount for p in all_payments if p.payment_type == 'Passport Fee')
                graduation_paid = sum(p.amount for p in all_payments if p.payment_type == 'Graduation Fee')
                total_paid = sum(p.amount for p in all_payments)
                student.payment_details = {
                    'passport': passport_paid,
                    'graduation': graduation_paid,
                    'total': total_paid
                }
                grad_total = graduation_paid
                if grad_total >= 1000:
                    total_graduated += 1
            
            report_data = students
    
    return render_template(
        'student_payment_status.html',
        sessions=sessions,
        classes=classes,
        report_data=report_data,
        selected_payment_type=selected_payment_type,
        selected_class=selected_class,
        selected_session=selected_session,
        total_enrolled=total_enrolled,
        total_graduated=total_graduated
    )

@main_routes.route('/add_student', methods=['GET', 'POST'])
@login_required
def add_student():
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        residence = request.form.get('residence')
        class_name = request.form.get('class_name')
        session = request.form.get('session')
        next_of_kin_name = request.form.get('next_of_kin_name', '')
        next_of_kin_relationship = request.form.get('next_of_kin_relationship', '')
        next_of_kin_phone = request.form.get('next_of_kin_phone', '')
        
        if not all([name, phone, residence, class_name, session]):
            flash('All fields are required!', 'error')
            return redirect(url_for('main.add_student'))
        
        student = Student(
            name=name,
            phone=phone,
            residence=residence,
            class_name=class_name,
            session=session,
            next_of_kin_name=next_of_kin_name,
            next_of_kin_relationship=next_of_kin_relationship,
            next_of_kin_phone=next_of_kin_phone
        )
        db.session.add(student)
        db.session.commit()

        # Auto-sync to Supabase
        try:
            from supabase_sync import sync_student_to_supabase
            sync_student_to_supabase(student)
        except Exception as e:
            # Log error but don't fail the operation
            print(f"Supabase sync error: {str(e)}")

        flash('Student added successfully!', 'success')
        return redirect(url_for('main.index'))
    
    return render_template('add_student.html')

@main_routes.route('/edit_student/<int:student_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    
    if request.method == 'POST':
        student.name = request.form.get('name')
        student.phone = request.form.get('phone')
        student.residence = request.form.get('residence')
        student.class_name = request.form.get('class_name')
        student.session = request.form.get('session')
        student.next_of_kin_name = request.form.get('next_of_kin_name', '')
        student.next_of_kin_relationship = request.form.get('next_of_kin_relationship', '')
        student.next_of_kin_phone = request.form.get('next_of_kin_phone', '')

        db.session.commit()

        # Auto-sync to Supabase
        try:
            from supabase_sync import sync_student_to_supabase
            sync_student_to_supabase(student)
        except Exception as e:
            print(f"Supabase sync error: {str(e)}")

        flash('Student updated successfully!', 'success')
        return redirect(url_for('main.index'))
    
    return render_template('edit_student.html', student=student)

@main_routes.route('/delete_student/<int:student_id>', methods=['POST'])
@login_required
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    db.session.delete(student)
    db.session.commit()
    flash('Student deleted successfully!', 'success')
    return redirect(url_for('main.index'))

@main_routes.route('/view_student/<int:student_id>')
@login_required
def view_student(student_id):
    student = Student.query.get_or_404(student_id)
    payments = Payment.query.filter_by(student_id=student_id).order_by(Payment.date.desc()).all()
    from models import ExamResult, Attendance
    from datetime import timedelta
    exam_results = ExamResult.query.filter_by(student_id=student_id).order_by(ExamResult.created_at.desc()).all()
    thirty_days_ago = datetime.now().date() - timedelta(days=30)
    recent_attendance = Attendance.query.filter(Attendance.student_id == student_id, Attendance.date >= thirty_days_ago).order_by(Attendance.date.desc()).all()
    return render_template('view_student.html', student=student, payments=payments, exam_results=exam_results, recent_attendance=recent_attendance)

@main_routes.route('/download_report', methods=['GET', 'POST'])
@login_required
def download_report():
    from models import SubAdmin
    from flask_login import current_user

    # Sub-admins cannot access payment reports
    if isinstance(current_user, SubAdmin):
        flash('Sub-admins do not have permission to access payment reports', 'error')
        return redirect(url_for('main.index'))
    selected_session = request.args.get('session') or request.form.get('session', '')
    selected_class = request.args.get('class_') or request.form.get('class', '')
    payment_type_param = request.args.get('payment_type') or request.form.get('payment_type', '')
    format_param = request.args.get('format') or request.form.get('format', 'csv')
    import csv
    from io import StringIO
    if not selected_session and not payment_type_param:
        return 'Session or payment_type required', 400
    # Compute payment details for download
    if payment_type_param and not selected_session:
        students = Student.query.filter_by(active=True).order_by(Student.class_name, Student.name).all()
    elif selected_class == "All Classes":
        students = Student.query.filter_by(session=selected_session).order_by(Student.class_name, Student.name).all()
    else:
        students = Student.query.filter_by(session=selected_session, class_name=selected_class).order_by(Student.name).all()
    
    for student in students:
        all_payments = student.payments
        student.passport_total = sum(p.amount for p in all_payments if p.payment_type == 'Passport Fee')
        student.graduation_total = sum(p.amount for p in all_payments if p.payment_type == 'Graduation Fee')
        student.transport_total = sum(p.amount for p in all_payments if p.payment_type == 'Transport Fee')
        student.total_paid = sum(p.amount for p in all_payments)
    
    # Headers for new columns
    headers = ["ID", "Name", "Phone", "Residence", "Class", "Passport", "Graduation", "Transport", "Total"]
    
    report_rows = []
    for student in students:
        report_rows.append([
            student.id,
            student.name,
            student.phone,
            student.residence or '',
            student.class_name,
            student.passport_total,
            student.graduation_total,
            student.transport_total,
            student.total_paid
        ])
    
    format_val = format_param
    if format_val == "excel":
        try:
            from openpyxl import Workbook
            from io import BytesIO
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            for row in report_rows:
                ws.append(row)
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment;filename=payment_report_{selected_session}_{selected_class or 'all'}.xlsx"}
            )
        except ImportError:
            return "Excel support requires openpyxl. Install with: pip install openpyxl", 500
    elif format_val == "pdf":
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
            styles = getSampleStyleSheet()
            header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=10, textColor=colors.white, alignment=1)
            cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)
            
            formatted_data = [[Paragraph(str(h), header_style) for h in headers]]
            for row in report_rows:
                formatted_data.append([Paragraph(str(cell), cell_style) for cell in row])
            
            col_widths = [40, 100, 80, 80, 60, 60, 60, 60, 60]
            table = Table(formatted_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            
            doc.build([table])
            buffer.seek(0)
            return Response(
                buffer.getvalue(),
                mimetype="application/pdf",
                headers={"Content-Disposition": f"attachment;filename=payment_report_{selected_session}_{selected_class or 'all'}.pdf"}
            )
        except ImportError:
            return "PDF support requires reportlab. Install with: pip install reportlab", 500
    else:
        si = StringIO()
        cw = csv.writer(si)
        cw.writerow(headers)
        cw.writerows(report_rows)
        return Response(
            si.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename=payment_report_{selected_session}_{selected_class or 'all'}.{format}"}
        )
    # Get students with the same sorting as the main view
    if selected_class == "All Classes":
        students = Student.query.filter_by(session=selected_session).order_by(Student.class_name, Student.name).all()
    else:
        students = Student.query.filter_by(session=selected_session, class_name=selected_class).order_by(Student.name).all()
    
    # Prepare report data with Graduation Fee status and amount
    report_rows = []
    for student in students:
        grad_payments = [p for p in student.payments if p.payment_type == 'Graduation Fee']
        grad_total = sum(p.amount for p in grad_payments)
        if grad_total >= 1000:
            grad_status = "Paid"
        elif grad_total > 0:
            grad_status = "Partial"
        else:
            grad_status = "Not Paid"
        passport_payments = [p for p in student.payments if p.payment_type == 'Passport Fee']
        passport_total = sum(p.amount for p in passport_payments)
        report_rows.append([
            student.id,  # Add student ID
            student.name,
            student.phone,
            student.residence,
            student.class_name,
            grad_total,
            passport_total
        ])
    headers = ["ID", "Name", "Phone", "Residence", "Class", "Graduation Fee Amount", "Passport Payment"]
    if format == "excel":
        from openpyxl import Workbook
        from io import BytesIO
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in report_rows:
            ws.append(row)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename=report_{selected_session}_{selected_class}.xlsx"}
        )
    elif format == "pdf":
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from io import BytesIO
        
        # Set up PDF with landscape orientation
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        
        # Define styles
        styles = getSampleStyleSheet()
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.white,
            alignment=1  # Center aligned
        )
        cell_style = ParagraphStyle(
            'Cell',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=1  # Center aligned
        )
        
        # Format data with proper styles
        formatted_data = []
        
        # Add headers with style
        header_row = [Paragraph(header, header_style) for header in headers]
        formatted_data.append(header_row)
        
        # Add data rows with style
        for row in report_rows:
            formatted_row = [Paragraph(str(cell), cell_style) for cell in row]
            formatted_data.append(formatted_row)
        
        # Create table with adjusted column widths
        col_widths = [40, 100, 80, 100, 60, 60, 50, 50]  # Adjusted widths for each column
        table = Table(formatted_data, colWidths=col_widths, repeatRows=1)  # Repeat header on each page
        
        # Apply table styles
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),  # Darker blue header
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#4F81BD')),
            ('LINEABOVE', (0, 1), (-1, -1), 0.5, colors.lightgrey),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ]))
        
        # Build the document with the table
        elements = [table]
        doc.build(elements)
        buffer.seek(0)
        return Response(
            buffer.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment;filename=report_{selected_session}_{selected_class}.pdf"}
        )
    else:
        import csv
        from io import StringIO
        si = StringIO()
        cw = csv.writer(si)
        cw.writerow(headers)
        cw.writerows(report_rows)
        output = si.getvalue()
        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename=report_{selected_session}_{selected_class}.csv"}
        )

@main_routes.route('/download_fee_list', methods=['POST'])
@login_required
def download_fee_list():
    fee_type = request.form.get('fee_type')
    if not fee_type or fee_type not in ['graduation', 'transport', 'both']:
        flash('Invalid fee type selected', 'error')
        return redirect(url_for('main.index'))
    
    from openpyxl import Workbook
    from io import BytesIO
    from sqlalchemy import func
    
    # Base query for paid students
    if fee_type == 'graduation':
        students_query = db.session.query(Student).\
            join(Payment, Student.id == Payment.student_id).\
            filter(Student.active == True, Payment.payment_type == 'Graduation Fee').\
            group_by(Student.id).\
            having(func.sum(Payment.amount) > 0).\
            order_by(Student.class_name, Student.name).all()
        filename = f"graduation_fee_students_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    elif fee_type == 'transport':
        students_query = db.session.query(Student).\
            join(Payment, Student.id == Payment.student_id).\
            filter(Student.active == True, Payment.payment_type == 'Transport Fee').\
            group_by(Student.id).\
            having(func.sum(Payment.amount) > 0).\
            order_by(Student.class_name, Student.name).all()
        filename = f"transport_fee_students_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    else:  # both
        students_query = db.session.query(Student).\
            join(Payment, Student.id == Payment.student_id).\
            filter(Student.active == True, Payment.payment_type.in_(['Graduation Fee', 'Transport Fee'])).\
            group_by(Student.id).\
            having(
                func.sum(db.case((Payment.payment_type == 'Graduation Fee', Payment.amount), else_=0)) > 0,
                func.sum(db.case((Payment.payment_type == 'Transport Fee', Payment.amount), else_=0)) > 0
            ).\
            order_by(Student.class_name, Student.name).all()
        filename = f"both_fees_students_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    # Compute totals for each student
    students = []
    for student in students_query:
        passport_total = sum(p.amount for p in student.payments if p.payment_type == 'Passport Fee')
        graduation_total = sum(p.amount for p in student.payments if p.payment_type == 'Graduation Fee')
        transport_total = sum(p.amount for p in student.payments if p.payment_type == 'Transport Fee')
        
        students.append({
            'id': student.id,
            'name': student.name,
            'phone': student.phone,
            'residence': student.residence or '',
            'class_name': student.class_name,
            'passport': passport_total,
            'graduation': graduation_total,
            'transport': transport_total
        })
    
    if not students:
        flash('No students found for the selected fee type.', 'info')
        return redirect(url_for('main.index'))
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Fee List"
    
    headers = ["ID", "Name", "Phone", "Residence", "Class", "Passport", "Graduation", "Transport"]
    ws.append(headers)
    
    for student in students:
        ws.append([
            student['id'],
            student['name'],
            student['phone'],
            student['residence'],
            student['class_name'],
            f"{student['passport']:.2f}" if student['passport'] > 0 else "0.00",
            f"{student['graduation']:.2f}" if student['graduation'] > 0 else "0.00",
            f"{student['transport']:.2f}" if student['transport'] > 0 else "0.00"
        ])
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@main_routes.route('/end_session', methods=['POST'])
@login_required
def end_session():
    try:
        session_number = request.form.get('session_number')
        if not session_number:
            flash('Please provide a session number', 'error')
            return redirect(url_for('main.index'))
            
        # Update all active students in the specified session to inactive
        updated_count = db.session.query(Student).filter(
            Student.session == str(session_number),
            Student.active == True
        ).update({'active': False}, synchronize_session=False)
        
        db.session.commit()
        flash(f'Successfully marked {updated_count} students from session {session_number} as inactive', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error ending session: {str(e)}', 'error')
    
    return redirect(url_for('main.index'))
